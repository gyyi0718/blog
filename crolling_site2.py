import os
import time
import zipfile
import re
import requests
import xml.etree.ElementTree as ET
from io import BytesIO
from collections import Counter
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from urllib.parse import urljoin
from openpyxl import load_workbook
import fitz  # PyMuPDF
import pdfplumber

SAVE_DIR = "금천구청_업무추진비_식비"
os.makedirs(SAVE_DIR, exist_ok=True)

headers = {"User-Agent": "Mozilla/5.0"}
options = Options()
options.add_argument("--headless")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

def extract_hwp(download_url):
    temp_list = []
    try:
        res = requests.get(download_url, headers=headers)
        res.raise_for_status()
        with zipfile.ZipFile(BytesIO(res.content)) as zip_ref:
            zip_ref.extractall("temp_hwpx")

        tree = ET.parse("temp_hwpx/Contents/section0.xml")
        root = tree.getroot()

        for elem in root.iter():
            if elem.tag.endswith("t") and elem.text:
                text = elem.text.strip()
                if any(kw in text for kw in ["사용처", "사용장소","장소"]):
                    if 1 < len(text) < 50:
                        temp_list.append(text)
    except Exception as e:
        print(f"❌ HWPX 파일 처리 실패: {e}")
    return temp_list

def extract_excel(download_url):
    temp_list = []
    try:
        res = requests.get(download_url, headers=headers)
        res.raise_for_status()
        with BytesIO(res.content) as excel_file:
            wb = load_workbook(filename=excel_file, read_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if isinstance(cell, str) and any(kw in cell for kw in ["사용처", "사용장소", "장소"]):
                            if 1 < len(cell) < 50:
                                temp_list.append(cell.strip())
    except Exception as e:
        print(f"❌ Excel 파일 처리 실패: {e}")
    return temp_list

# 키워드 필터링을 제거하여 모든 텍스트를 대상으로 추출
def extract_from_pdf(download_url):
    temp_list = []
    try:
        pdf_res = requests.get(download_url, headers=headers, stream=True, timeout=10)
        pdf_res.raise_for_status()
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
            for chunk in pdf_res.iter_content(8192):
                tmp_pdf.write(chunk)
            tmp_pdf_path = tmp_pdf.name
        try:
            with pdfplumber.open(tmp_pdf_path) as pdf_file:
                with pdfplumber.open(pdf_file) as pdf:
                    for page in pdf.pages:
                        try:
                            tables = page.extract_tables()
                            for table in tables:
                                df = pd.DataFrame(table)
                                df.columns = df.iloc[0]
                                df = df[1:]
                                df.columns = df.columns.astype(str)
                                for col in df.columns:
                                    if any(keyword in col for keyword in ["사용처", "사용장소","장소"]):
                                        for val in df[col].dropna():
                                            val_clean = str(val).strip()
                                            if 1 < len(val_clean) < 50:
                                                temp_list.append(val_clean)
                                print("pdf : ",df.head(5))
                        except Exception as page_err:
                            print(f"⚠️ PDF 페이지 처리 실패: {page_err}")
        except Exception as pdf_err:
            print(f"❌ PDF 열기 실패: {pdf_err}")

    except Exception as e:
        print(f"❌ 파일 다운로드 또는 처리 실패: {download_url}, {e}")

    return temp_list

def get_attachment_urls(detail_url):
    driver.get(detail_url)
    time.sleep(1)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    attachment_tags = soup.select("div.bo_file a")
    download_links = []
    for a in attachment_tags:
        href = a.get("href")
        if href:
            full_link = urljoin("https://www.geumcheon.go.kr", href)
            download_links.append(full_link)
    return download_links

def site_geumcheon_list(base_list_url):
    restaurant_list = []
    for page in range(1, 3):  # 필요한 페이지 수 조정
        driver.get(f"{base_list_url}&page={page}")
        time.sleep(1)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        rows = soup.select("table.tbl01 tbody tr")

        for row in rows:
            a_tag = row.select_one("td.p-subject a")
            if not a_tag:
                continue
            title = a_tag.get_text(strip=True)
            href = a_tag.get("href")
            detail_url = urljoin("https://www.geumcheon.go.kr", href)

            print(f"🔎 게시글: {title} → {detail_url}")
            try:
                file_urls = get_attachment_urls(detail_url)
                for file_url in file_urls:
                    if file_url.endswith(".hwpx"):
                        extracted = extract_hwp(file_url)
                        restaurant_list.extend(extracted)
                    elif file_url.endswith(".xlsx"):
                        extracted = extract_excel(file_url)
                        restaurant_list.extend(extracted)
                    elif file_url.endswith(".pdf"):
                        extracted = extract_from_pdf(file_url)
                        restaurant_list.extend(extracted)
            except Exception as e:
                print(f"❌ 첨부파일 처리 실패: {e}")

    # ✅ 결과 저장
    restaurant_counter = Counter(restaurant_list)
    final_df = pd.DataFrame(restaurant_counter.items(), columns=["사용장소", "사용횟수"])
    final_df = final_df.sort_values(by="사용횟수", ascending=False)
    final_df.to_csv(os.path.join(SAVE_DIR, "금천구청_식당_사용빈도.csv"), index=False, encoding="utf-8-sig")
    print("✅ 식당 사용 빈도 분석 결과 저장 완료")

if __name__ == "__main__":
    base_list_url = "https://www.geumcheon.go.kr/portal/selectBbsNttList.do?bbsNo=86&key=269"
    site_geumcheon_list(base_list_url)
    driver.quit()
