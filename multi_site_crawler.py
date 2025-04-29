import os
import time
import shutil
import zipfile
import requests
import pandas as pd
import pdfplumber
import xml.etree.ElementTree as ET
import re
from typing import List, Tuple
from pathlib import Path
from collections import Counter
from urllib.parse import urljoin, unquote
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
import csv
import sys
import zipfile
import shutil
import xml.etree.ElementTree as ET
import pdfplumber
from urllib.parse import unquote

os.system('')                # Virtual Terminal 활성화 (Windows 10 이상)
sys.stdout.reconfigure(encoding='utf-8')



class SiteConfig:
    """
    Configuration for a single site to crawl.
    - list_url_template: URL with a {page} placeholder for pagination.
    - detail_base: Base URL to resolve relative detail links.
    - row_selector: CSS selector for rows on the list page.
    - title_selector: CSS selector within a row to get the title/link.
    - attach_selector: CSS selector on detail page for attachment links.
    - pages: iterable of page numbers to crawl.
    - output_csv: filename for final report.
    """
    def __init__(
        self,
        name: str,
        list_url_template: str,
        detail_base: str,
        row_selector: str,
        title_selector: str,
        attach_selector: str,
        pages=range(1, 6),
        output_csv="usage_report.csv"
    ):
        self.name = name
        self.list_url_template = list_url_template
        self.detail_base = detail_base
        self.row_selector = row_selector
        self.title_selector = title_selector
        self.attach_selector = attach_selector
        self.pages = pages
        self.output_csv = output_csv


class SiteParser:
    """
    Base parser: override methods for site-specific extraction rules.
    """
    def __init__(self, config: SiteConfig):
        self.config = config

    def extract_post_links(self, html: str):
        raise NotImplementedError

    def extract_attachment_links(self, html: str):
        raise NotImplementedError


class DefaultParser(SiteParser):
    """
    Default parser assumes similar structure across sites.
    """
    def extract_post_links(self, html: str):
        soup = BeautifulSoup(html, 'html.parser')
        rows = soup.select(self.config.row_selector)
        links = []
        for row in rows:
            a = row.select_one(self.config.title_selector)
            if a and a.get('href'):
                links.append(urljoin(self.config.detail_base, a['href']))
        return links

    def extract_attachment_links(self, html: str):
        soup = BeautifulSoup(html, 'html.parser')
        attachments = []
        for a in soup.select(self.config.attach_selector):
            href = a.get('href')
            name = a.get_text(strip=True) or Path(href).name
            if href:
                attachments.append((urljoin(self.config.detail_base, href), name))
        return attachments

class GeumcheonParser(SiteParser):
    row_selector = 'table.p-table.simple tbody tr'
    title_selector = 'td.p-subject a'
    attach_selector = 'a.p-attach__link'

    def extract_post_links(self, html: str) -> List[str]:
        soup = BeautifulSoup(html, "html.parser")
        rows = soup.select(self.row_selector)
        links = []
        for row in rows:
            a_tag = row.select_one(self.title_selector)
            if not (a_tag and a_tag.get('href')):
                continue
            # 선행 슬래시 제거
            href = a_tag['href']
            # 기본 detail_base: "https://www.geumcheon.go.kr"
            # href 가 "selectBbsNttView.do?..." 형태라면 portal/ 을 붙여서 절대경로 생성

            if href.startswith('/'):
                full_url = urljoin(self.config.detail_base, href)
            else:
                full_url = urljoin(self.config.detail_base + '/portal/', href.lstrip('/'))
            links.append(full_url)
        return links

    def extract_attachment_links(self, html: str) -> List[Tuple[str, str]]:
        soup = BeautifulSoup(html, "html.parser")
        files = []
        for a in soup.select(self.attach_selector):
            href = a.get('href')
            name = a.get_text(strip=True)
            if not href:
                continue
            href = href.lstrip('/')
            full_url = f"{self.config.detail_base}/{href}"
            files.append((full_url, name))
        return files
class MapoParser(SiteParser):
    def __init__(self, config: SiteConfig):
        super().__init__(config)
        self.detail_base = "https://www.mapo.go.kr"
        # 게시판 리스트에서 <tr> 을 골라내는 셀렉터
        self.row_selector   = 'ul.bbs_list li'       # 실제 list 페이지 구조에 맞춰 주세요
        self.title_selector = 'a.bbs_tit'                  # 리스트에서 게시글 링크 골라내기
        self.attach_selector = 'div.bbs_view_file a.file_name'

    def extract_post_links(self, html: str) -> List[str]:
        """리스트 페이지에서 /site/main/board/expense/숫자… 링크만 골라서 반환"""
        soup = BeautifulSoup(html, "html.parser")
        links: List[str] = []
        # href가 '/site/main/board/expense/숫자' 형태인 <a> 태그만 선택
        for a in soup.find_all("a", href=re.compile(r"^/site/main/board/expense/\d+")):
            href = a["href"]
            full = urljoin(self.config.detail_base, href)
            links.append(full)
        return links

    def extract_attachment_links(self, html: str) -> List[Tuple[str, str]]:
        """디테일 페이지에서 class="file_name"인 첨부파일만 추출"""
        soup = BeautifulSoup(html, "html.parser")
        files: List[Tuple[str, str]] = []
        for a in soup.select("a.file_name"):
            href = a.get("href")
            name = a.get_text(strip=True)
            if not href:
                continue
            full = urljoin(self.config.detail_base, href)
            files.append((full, name))
        return files

class GwangjuParser(SiteParser):
    def extract_post_links(self, html: str) -> List[str]:
        soup = BeautifulSoup(html, "html.parser")
        rows = soup.select("table.bod_list tbody tr")
        links = []
        for row in rows:
            a = row.select_one("td.list_tit a")
            if a:
                onclick = a.get("onclick", "")
                bidx = self.extract_bidx_from_onclick(onclick)
                if bidx:
                    detail_url = f"https://www.gjcity.go.kr/portal/bbs/view.do?bIdx={bidx}&ptIdx=53&mId=0311000000"
                    links.append(detail_url)
        return links

    def extract_attachment_links(self, html: str) -> List[Tuple[str, str]]:
        # 이건 selenium으로 처리할 거라 여기선 사용 X
        return []

    def extract_bidx_from_onclick(self, onclick_str: str) -> str:
        try:
            return onclick_str.split("'")[3]
        except IndexError:
            return ""

    def selenium_download_attachments(self, driver, download_dir: Path) -> List[Path]:
        time.sleep(1)
        files = []
        soup = BeautifulSoup(driver.page_source, "html.parser")
        file_links = soup.select("ul.file_list li a.download")
        for a in file_links:
            onclick = a.get("onclick", "")
            if "download" in onclick:
                try:
                    parts = onclick.split("'")
                    atchFileId = parts[1]
                    fileSn = parts[3]
                    download_url = f"https://www.gjcity.go.kr/common/file/download.do?atchFileId={atchFileId}&fileSn={fileSn}"
                    filename = a.get_text(strip=True)
                    filename = re.sub(r'\[\d+(\.\d+)?MB\]', '', filename)  # 중간에 붙은 경우
                    filename = re.sub(r'(\.\w+)\[\d+(\.\d+)?MB\]$', r'\1', filename)  # 확장자 뒤에 붙은 경우

                    out_path = download_dir / filename

                    resp = requests.get(download_url)
                    with open(out_path, "wb") as f:
                        f.write(resp.content)
                    print(f"📥 Downloaded: {out_path}")
                    files.append(out_path)
                except Exception as e:
                    print(f"❌ Error downloading: {e}")
        return files


class Crawler:
    """
    Core crawler: fetch list pages, detail pages, download attachments, extract data.
    """
    def __init__(self, config: SiteConfig, parser: SiteParser, download_dir: str = 'temp'):
        self.config = config
        self.parser = parser
        self.download_dir = Path(download_dir) / config.name
        self.download_dir.mkdir(parents=True, exist_ok=True)
        self.counter = Counter()

        chrome_opts = Options()
        chrome_opts.add_argument('--headless')
        self.driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=chrome_opts
        )

    def fetch_html(self, url: str) -> str:
        try:
            response = requests.get(url)
            response.encoding = 'utf-8'
            response.raise_for_status()
            return response.text
        except Exception as e:
            print(f"❌ 오류 발생: {url} - {e}")
            return ""

    def download_file(self, url: str, name_hint: str = "") -> Path:
        resp = requests.get(url, stream=True, timeout=10)
        resp.raise_for_status()

        # 1) Content-Disposition에서 파일명 추출
        cd = resp.headers.get("Content-Disposition", "")
        fname = None
        if "filename=" in cd:
            raw = cd.split("filename=")[-1].strip().strip('"')
            raw_bytes = raw.encode("latin1", errors="ignore")
            for enc in ("utf-8", "cp949"):
                try:
                    decoded = raw_bytes.decode(enc)
                    fname = decoded
                    break
                except Exception:
                    continue

        # 2) 힌트나 URL에서 추출
        if not fname and name_hint:
            fname = name_hint
        if not fname:
            fname = Path(url).name


        # 3) 용량 정보 제거 (예: "파일명.pdf[0.16MB]")
        fname = re.sub(r'\[\d+(\.\d+)?MB\]$', '', fname)

        # 4) 특수문자, 연속 공백 등 제거
        fname = re.sub(r'[\x00-\x1f\x7f]', "", fname)
        fname = re.sub(r"\s+", " ", fname).strip()
        fname = re.sub(r'[\\/:*?"<>|]+', "_", fname)
        fname = re.sub(r'[;_]+$', "", fname)
        # 3) 용량 정보 제거: [...MB] 형식 제거 (.pdf[0.16MB], [0.16MB].pdf 모두 포함)
        fname = re.sub(r'\[\d+(\.\d+)?MB\]', '', fname)  # 중간에 붙은 경우
        fname = re.sub(r'(\.\w+)\[\d+(\.\d+)?MB\]$', r'\1', fname)  # 확장자 뒤에 붙은 경우

        out_path = self.download_dir / fname
        with open(out_path, "wb") as f:
            for chunk in resp.iter_content(8192):
                f.write(chunk)

        size_mb = out_path.stat().st_size / 1024 / 1024
        print(f"📥 Downloaded: {out_path}[{size_mb:.2f}MB]")
        return out_path
    def clean_name(self, raw: str) -> str:
        """
        1) 큰따옴표 제거
        2) 지역명(서울시, 서울, 금천구, 금천, 부산시, 부산, 해운대구, 해운대, 수영구, 수영) 전부 제거
        3) 쉼표나 백슬래시 뒤 정보 삭제
        4) 소괄호 및 내부 텍스트 삭제
        5) 온점(.) 제거
        6) 모든 공백(스페이스, 탭, 줄바꿈) 제거
        """
        s = raw.replace('"', '').strip()

        # (2) 지역명 제거 — \b 대신 그냥 연속 패턴으로
        s = re.sub(r'(서울시|서울|금천구|금천|부산시|부산|해운대구|해운대|수영구|수영)', '', s)

        # (3) 쉼표나 백슬래시 뒤 정보 삭제
        s = re.sub(r'[,\\].*$', '', s)

        # (4) 소괄호 안 텍스트 삭제
        s = re.sub(r'\([^)]*\)', '', s)

        # (5) 온점 제거
        s = s.replace('.', '')

        # (6) 모든 공백 제거
        s = re.sub(r'\s+', '', s)

        return s
    def normalize_restaurant_name(name: str, address: str, region: str) -> str:
        """식당명 문자열을 정규화하여 불필요한 주소/분점/부가 정보를 제거하고 순수 식당명으로 반환"""
        # 1. 지역명 포함 주소 부분 제거 (예: "서울시 금천구 ..." 제거)
        idx = -1
        # 지역별 도시 이름 패턴 설정
        if region.endswith('구'):
            if region in ['해운대구', '수영구']:
                city_keywords = ['부산광역시', '부산시', '부산']
            elif region == '금천구':
                city_keywords = ['서울특별시', '서울시', '서울']
            else:
                city_keywords = ['서울특별시', '서울시', '서울', '부산광역시', '부산시', '부산']
            # "도시 지역" 패턴 탐색
            for city in city_keywords:
                phrase = f"{city} {region}"
                if phrase in name:
                    idx = name.index(phrase)
                    break
        # 도시명을 포함하지 않고 지역명만 있는 경우도 처리
        if idx == -1:
            if region + ' ' in name:  # 지역명 뒤에 공백이 있는 경우
                idx = name.index(region + ' ')
            elif name.endswith(region):  # 지역명으로 이름이 끝나는 경우
                idx = name.index(region)
        # 식당명과 공백으로 구분된 지역명일 때만 잘라냄
        if idx != -1:
            if idx == 0 or name[idx - 1].isspace():
                name = name[:idx]
        # 2. 괄호로 된 부가정보 제거 (대표자명 등)
        name = re.sub(r'\([^)]*\)', '', name)
        # 3. 층수 정보 제거 (예: "1층" 등의 패턴)
        name = re.sub(r'\s*\d+층$', '', name)
        # 4. 분점 표시 제거 (숫자 및 "관"/"호점" 또는 "본점"으로 끝나는 경우)
        name = re.sub(r'\s*(?:\d+\s*(?:관|호점)?|본점)$', '', name)
        # 5. 불필요한 공백 정리 (양쪽 공백 제거 및 중복 공백 축약)


        name = ' '.join(name.split())
        return name

    def reaggregate_csv(self, input_csv: str, output_csv: str):
        """
        input_csv 에 있는 'place, count' 데이터를
        주소의 쉼표는 무시하고 맨 마지막 쉼표만 기준으로 분리해
        place 를 정제(clean_name)한 뒤 count 를 합산해
        output_csv 로 다시 저장합니다.
        """
        ctr = Counter()
        with open(input_csv, encoding='utf-8-sig') as f:
            next(f)  # 헤더 건너뛰기
            for line in f:
                line = line.strip()
                if not line:
                    continue
                # 뒤에서 한 번만 나눠서, 왼쪽은 place, 오른쪽은 count
                place, cnt = line.rsplit(',', 1)
                place = self.clean_name(place)
                try:
                    num = int(cnt)
                except ValueError:
                    continue
                ctr[place] += num

        # 결과를 output_csv 로 저장
        with open(output_csv, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['place', 'count'])  # 헤더
            for place, count in ctr.most_common():
                writer.writerow([place, count])

        print(f"✅ 재집계 완료: {output_csv}")

    def extract_from_file(self, path: Path):
        def safe_text(x):
            if isinstance(x, str):
                return x.replace('\x00', '').replace('\ufffd', '').strip()
            return str(x).strip()

        ext = path.suffix.lower()
        try:
            if ext in ('.xls', '.xlsx'):
                wb = openpyxl.load_workbook(path, data_only=True)
                visible_sheets = [ws for ws in wb.worksheets if ws.sheet_state == 'visible']

                if not visible_sheets:
                    print(f"⚠️ 보이는 시트 없음: {path.name}")
                    return

                ws = visible_sheets[0]  # 첫 번째 보이는 시트만 읽기

                data = []
                for row in ws.iter_rows(values_only=True):
                    if all(cell is None for cell in row):
                        continue
                    data.append([safe_text(cell) for cell in row])

                df = pd.DataFrame(data)
                print(f"✅ 열람 성공: {path.name}")

                # 컬럼 직접 탐색
                found = False
                for idx, row in df.iterrows():
                    if any(any(k in str(cell) for k in ['사용처', '사용장소', '장소']) for cell in row):
                        df.columns = row
                        df = df.drop(index=range(idx + 1))
                        print(f"📌 컬럼 (row {idx}): {df.columns.tolist()}")
                        found = True
                        break

                if not found:
                    print(f"⚠️ 사용처/사용장소 컬럼 없음: {path.name}")
                    return

                cols = [c for c in df.columns if any(k in safe_text(c) for k in ['사용처', '사용장소', '장소'])]

                for col in cols:
                    for v in df[col].dropna().astype(str):
                        name = self.clean_name(v)
                        print(f"📍 추출된 이름: {name}")
                        if 1 < len(name) < 50:
                            self.counter[name] += 1

            elif ext == '.pdf':
                if path.stat().st_size < 1024:
                    print(f"⚠️ {path} 파일 사이즈 너무 작음 (1KB 이하)")
                    return
                with pdfplumber.open(path) as pdf:
                    for pg in pdf.pages:
                        for tbl in pg.extract_tables() or []:
                            df = pd.DataFrame(tbl[1:], columns=[safe_text(c) for c in tbl[0]])
                            for col in df.columns:
                                if any(k in safe_text(col) for k in ['사용처', '사용장소', '장소']):
                                    for v in df[col].dropna().astype(str):
                                        name = self.clean_name(v)
                                        if 1 < len(name) < 50:
                                            self.counter[name] += 1

            elif ext == '.hwpx':
                tmp = self.download_dir / '_hwpx'
                with zipfile.ZipFile(path) as zp:
                    zp.extractall(tmp)
                xmlf = tmp / 'Contents' / 'section0.xml'
                tree = ET.parse(xmlf)
                for t in tree.iter():
                    if t.tag.endswith('}t') and t.text:
                        text = safe_text(t.text)
                        if any(k in text for k in ['사용처', '사용장소', '장소']):
                            name = self.clean_name(text)
                            if 1 < len(name) < 50:
                                self.counter[name] += 1
                shutil.rmtree(tmp)

            else:
                print(f"Unsupported format: {path}")

        except Exception as e:
            print(f"Error parsing {path.name} ({ext}): {e}")

    def save_to_csv(self, output_path: str):
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['place', 'count'])
            for place, count in sorted(self.counter.items(), key=lambda x: -x[1]):
                writer.writerow([self.safe_text(place), count])
        print(f"✅ CSV 저장 완료: {output_path}")
    def run(self, max_pages: int = None):
        pages = self.config.pages if max_pages is None else range(1, max_pages+1)
        for p in pages:
            try:
                list_url = self.config.list_url_template.format(page=p)
                print(f"Crawling list page: {list_url}")
                html = self.fetch_html(list_url)
                if not html: continue

                links = self.parser.extract_post_links(html)
                for post in links:
                    try:
                        print(f"Visiting post: {post}")
                        self.driver.get(post)
                        time.sleep(1)
                        page_html = self.driver.page_source
                        attaches = self.parser.extract_attachment_links(page_html)
                        for url, hint in attaches:
                            try:
                                fpath = self.download_file(url, hint)
                                if fpath:
                                    try:
                                        self.extract_from_file(fpath)
                                    finally:
                                        # 파일 읽든 실패하든 무조건 삭제
                                        fpath.unlink(missing_ok=True)  # Path 객체라 unlink()
                            except Exception as e:
                                print(f"Error file Pathe {fpath} : {e}")
                    except Exception as e:
                        print(f"Error Post : {e}")
            except Exception as e:
                print(f"Error page : {e}")
        # save report
        df = pd.DataFrame(self.counter.most_common(), columns=['place','count'])
        df.to_csv(
            self.config.output_csv,
            index=False,
            encoding='utf-8-sig',
            quoting=csv.QUOTE_NONE,  # 따옴표 감싸기 없이
            escapechar='\\'  # 이스케이프 문자 지정
        )
        print(f"Report saved to {self.config.output_csv}")

    def cleanup(self):
        if self.download_dir.exists():
            shutil.rmtree(self.download_dir)
            print(f"Cleaned up download dir: {self.download_dir}")

class GwangjuCrawler:
    def __init__(self, download_dir: str = "temp/광주시청"):
        self.download_dir = Path(download_dir)
        self.download_dir.mkdir(parents=True, exist_ok=True)

        chrome_options = Options()
        prefs = {
            "download.default_directory": str(self.download_dir.resolve()),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        chrome_options.add_argument("--headless=new")
        chrome_options.add_argument("--disable-gpu")
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        self.download_dir = Path("temp") / "광주시청"
        self.download_dir.mkdir(parents=True, exist_ok=True)
        self.counter = Counter()

    def run(self, max_page=1):
        for page in range(1, max_page + 1):
            list_url = f"https://www.gjcity.go.kr/portal/bbs/list.do?ptIdx=53&mId=0311000000&page={page}"
            print(f"Crawling list page: {list_url}")
            self.driver.get(list_url)
            soup = BeautifulSoup(self.driver.page_source, "html.parser")
            rows = soup.select("table.bod_list tbody tr")

            for row in rows:
                a_tag = row.select_one("td.list_tit a")
                onclick = a_tag.get("onclick", "")
                bidx = self.extract_bidx_from_onclick(onclick)
                if not bidx:
                    continue
                post_url = f"https://www.gjcity.go.kr/portal/bbs/view.do?bIdx={bidx}&ptIdx=53&mId=0311000000"
                print(f"Visiting post: {post_url}")
                self.driver.get(post_url)
                time.sleep(1.5)

                self.download_attachments_from_page()


    def extract_bidx_from_onclick(self, onclick: str) -> str:
        try:
            return onclick.split("'")[3]
        except IndexError:
            return None

    def download_attachments_from_page(self):
        file_links = self.driver.find_elements(By.CSS_SELECTOR, "ul.file_list li a.download")
        for a in file_links:
            try:
                filename = a.text.strip()
                print(f"📎 클릭 중: {filename}")
                before_files = set(self.download_dir.glob("*"))
                a.click()
                self.wait_for_download(before_files)
            except Exception as e:
                print(f"❌ 다운로드 실패: {e}")

    def wait_for_download(self, before_files: set, timeout=15):
        for _ in range(timeout):
            after_files = set(self.download_dir.glob("*"))
            new_files = after_files - before_files
            if new_files:
                for file in new_files:
                    print(f"✅ 다운로드 완료: {file.name}")
                return
            time.sleep(1)
        print("❌ 다운로드 시간 초과")

    def close(self):
        self.driver.quit()

def run_geumcheon():
    config = [SiteConfig(
        name="금천구청",
        list_url_template=(
            "https://www.geumcheon.go.kr/portal/selectBbsNttList.do"
            "?bbsNo=86&key=269&id=&searchCtgry=&pageUnit=10"
            "&searchCnd=all&searchKrwd=&integrDeptCode="
            "&searchDeptCode=&pageIndex={page}"
        ),
        # '/portal'까지 꼭 포함
        detail_base="https://www.geumcheon.go.kr/portal/",
        row_selector="table.p-table.simple tbody tr",
        title_selector="td.p-subject a",
        attach_selector="a.p-attach__link",
        pages=range(18, 83),
        output_csv="금천구청_업무추진비.csv"
    ),
        SiteConfig(
            name="구로구청",
            list_url_template=(
                "https://www.guro.go.kr/www/selectBbsNttList.do?"
                "bbsNo=655&pageUnit=10&key=1732&pageIndex={page}"
            ),
            # '/portal'까지 꼭 포함
            detail_base="https://www.guro.go.kr/www/",
            row_selector="table.p-table.simple tbody tr",
            title_selector="td.p-subject a",
            attach_selector="a.p-attach__link",
            pages=range(20, 100),
            output_csv="구로구청_업무추진비.csv"
        ),
    ]
    guro_config = config[1]
    # 18, 83 -> 2024.01~12 까지의 page 범위
    parser = DefaultParser(guro_config)
    crawler = Crawler(guro_config, parser)
    crawler.run(max_pages=None)
    crawler.cleanup()
    # crawler.reaggregate_csv('구로청_식당_사용빈도.csv', '구로청_식당_사용빈도_정제.csv')
    # crawler.reaggregate_csv('금천구청_업무추진비.csv', '금천구청_업무추진비_정제2.csv')

def run_mapo():
    mapo_config = SiteConfig(
        name='마포구청',
        list_url_template=(
            'https://www.mapo.go.kr/site/main/board/expense/list'
            '?cp={page}&sortOrder=BA_REGDATE'
            '&sortDirection=DESC'
            '&listType=list'
            '&bcId=expense'
            '&baUse=true'
        ),
        detail_base='https://www.mapo.go.kr/site/main/',
        row_selector='',  # DefaultParser는 사용 안 함
        title_selector='',
        attach_selector='',
        pages=range(20, 91),
        output_csv='마포구청_업무추진비.csv'
    )
    # 20, 91
    parser = MapoParser(mapo_config)
    crawler = Crawler(mapo_config, parser)
    crawler.run(max_pages=100)
    crawler.download_dir


def run_gwangju():
    config = SiteConfig(
        name="광주시청",
        list_url_template="https://www.gjcity.go.kr/portal/bbs/list.do?ptIdx=53&mId=0311000000&page={page}",
        detail_base="https://www.gjcity.go.kr",
        row_selector="table.bod_list tbody tr",
        title_selector="td.list_tit a",
        attach_selector="ul.file_list li a.download",
        pages=range(15, 73),
        output_csv="광주시청_업무추진비.csv"
    )
    parser = GwangjuParser(config)
    crawler = Crawler(config, parser)

    for page in config.pages:
        list_url = config.list_url_template.format(page=page)
        html = crawler.fetch_html(list_url)
        if not html:
            continue
        posts = parser.extract_post_links(html)
        for url in posts:
            print(f"Visiting post: {url}")
            print(f"Visiting post: {url}")
            crawler.driver.get(url)
            files = parser.selenium_download_attachments(crawler.driver, crawler.download_dir)
            for f in files:
                crawler.extract_from_file(f)
                f.unlink(missing_ok=True)  # Path 객체라 unlink()


    # save
    df = pd.DataFrame(crawler.counter.most_common(), columns=['place', 'count'])
    df.to_csv(config.output_csv, index=False, encoding='utf-8-sig')
    print(f"✅ CSV 저장 완료: {config.output_csv}")


if __name__ == '__main__':
    run_gwangju()
